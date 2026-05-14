#!/usr/bin/env python3
"""
Convert CommCare XForm XML files to KoboToolbox XLSForm workbooks.

Runtime behavior is controlled by environment variables. Put local settings
and secrets in `.env`; see `.env.example` for the supported keys.
"""

from __future__ import annotations

import argparse
import html
import io
import os
import re
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from tqdm import tqdm


PROJECT_ROOT = Path(__file__).resolve().parent


def _load_dotenv(path: Path) -> None:
    """Load simple KEY=value pairs from .env without adding a dependency."""
    if not path.exists():
        return
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def _env_bool(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None or value.strip() == "":
        return default
    normalized = value.strip().lower()
    if normalized in {"1", "true", "yes", "y", "on"}:
        return True
    if normalized in {"0", "false", "no", "n", "off"}:
        return False
    raise ValueError(
        f"Invalid boolean value for {name}: {value!r}. "
        "Use true/false, yes/no, on/off, or 1/0."
    )


def _env_int(name: str, default: int) -> int:
    value = os.getenv(name)
    if value is None or value.strip() == "":
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise ValueError(f"Invalid integer value for {name}: {value!r}") from exc


_load_dotenv(PROJECT_ROOT / ".env")

# Runtime settings. Values come from `.env` or process environment; defaults
# keep local dry/folder workflows usable when a key is omitted.
XML_INPUT_FOLDER = Path(os.getenv("XML_INPUT_FOLDER", PROJECT_ROOT / "XML_INPUT_FOLDER"))
XLSFORM_OUTPUT_FOLDER = Path(os.getenv("XLSFORM_OUTPUT_FOLDER", PROJECT_ROOT / "XLS_OUTPUT"))
SAVE_XLSFORMS_LOCALLY = _env_bool("SAVE_XLSFORMS_LOCALLY", True)

UPLOAD_TO_KOBO = _env_bool("UPLOAD_TO_KOBO", False)
KOBO_API_TOKEN = os.getenv("KOBO_API_TOKEN", "")
KOBO_SERVER_URL = os.getenv("KOBO_SERVER_URL", "https://eu.kobotoolbox.org")
KOBO_DEPLOY = _env_bool("KOBO_DEPLOY", False)

COMMCARE_FETCH = _env_bool("COMMCARE_FETCH", False)
COMMCARE_DOMAIN = os.getenv("COMMCARE_DOMAIN", "")
COMMCARE_USER = os.getenv("COMMCARE_USER", "")
COMMCARE_TOKEN = os.getenv("COMMCARE_TOKEN", "")
COMMCARE_LIMIT = _env_int("COMMCARE_LIMIT", 0)
COMMCARE_BASE_URL = os.getenv("COMMCARE_BASE_URL", "https://www.commcarehq.org")

XFORM_NS = "http://www.w3.org/2002/xforms"
HTML_NS = "http://www.w3.org/1999/xhtml"
VALID_XLSFORM_NAME = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
VALID_CHOICE_NAME = re.compile(r"^[^\s]+$")


def _commcare_session(user: str, token: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "Authorization": f"ApiKey {user}:{token}",
        "Accept": "application/json",
    })
    return session


def _raise_commcare_error(resp: requests.Response, context: str) -> None:
    try:
        message = resp.json().get("error", resp.text)
    except ValueError:
        message = resp.text[:500]

    if resp.status_code == 401 and "subscription" in message.lower():
        raise PermissionError(
            f"{context} failed: CommCare returned 401 - {message}. "
            "This endpoint requires a CommCare plan with API access."
        )
    if resp.status_code == 401:
        raise PermissionError(f"{context} failed: CommCare returned 401 - {message}")
    if resp.status_code == 403:
        raise PermissionError(f"{context} failed: CommCare returned 403 - {message}")
    if resp.status_code == 404:
        raise FileNotFoundError(
            f"{context} failed: CommCare returned 404. Check COMMCARE_BASE_URL and COMMCARE_DOMAIN."
        )

    resp.raise_for_status()


def _list_commcare_apps(session: requests.Session, domain: str, limit: int = 0) -> list[dict]:
    url = f"{COMMCARE_BASE_URL.rstrip('/')}/a/{domain}/api/application/v1/"
    params = {"format": "json", "limit": 100, "offset": 0}
    apps: list[dict] = []
    print("  Checking CommCare Application Structure API...")

    while True:
        resp = session.get(url, params=params, timeout=30)
        if not resp.ok:
            _raise_commcare_error(resp, "CommCare application fetch")
        data = resp.json()
        apps.extend(data.get("objects", []))
        if not data.get("meta", {}).get("next") or (limit and len(apps) >= limit):
            break
        params["offset"] += params["limit"]
        time.sleep(0.2)

    return apps[:limit] if limit else apps


def _list_commcare_fixture_rows(session: requests.Session, domain: str, fixture_type: str) -> list[dict[str, Any]]:
    url = f"{COMMCARE_BASE_URL.rstrip('/')}/a/{domain}/api/fixture/v1/"
    params = {"fixture_type": fixture_type, "limit": 1000, "offset": 0}
    rows: list[dict[str, Any]] = []

    while True:
        resp = session.get(url, params=params, timeout=60)
        if not resp.ok:
            _raise_commcare_error(resp, f"CommCare fixture fetch for {fixture_type}")
        data = resp.json()
        rows.extend(data.get("objects", []))
        if not data.get("meta", {}).get("next"):
            break
        params["offset"] += params["limit"]
        time.sleep(0.2)

    return rows


def _collect_fixture_types(apps: list[dict[str, Any]]) -> list[str]:
    fixture_types: list[str] = []
    seen: set[str] = set()
    for app in apps:
        for module in app.get("modules", []):
            for form in module.get("forms", []):
                for question in form.get("questions", []):
                    data_source = question.get("data_source") or {}
                    fixture_type = str(data_source.get("instance_id") or "")
                    if not fixture_type:
                        match = re.search(r"item-list:([^'\"/]+)", str(data_source.get("instance_ref") or ""))
                        fixture_type = match.group(1) if match else ""
                    if fixture_type and fixture_type not in seen:
                        seen.add(fixture_type)
                        fixture_types.append(fixture_type)
    return fixture_types


def _strip_ns(tag: str) -> str:
    return re.sub(r"\{[^}]+\}", "", tag)


def _normalize_ref(ref: str) -> str:
    ref = (ref or "").strip()
    if not ref:
        return ""
    ref = re.sub(r"^#form/?", "/data/", ref)
    if not ref.startswith("/"):
        ref = f"/data/{ref}"
    return re.sub(r"/+", "/", ref.rstrip("/"))


def _ref_leaf(ref: str) -> str:
    ref = _normalize_ref(ref)
    return ref.rsplit("/", 1)[-1] if ref else ""


def _safe_identifier(value: str, fallback: str, max_len: int = 64) -> str:
    value = html.unescape(str(value or "")).strip()
    value = re.sub(r"[^A-Za-z0-9_]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    if not value:
        value = fallback
    if not re.match(r"^[A-Za-z_]", value):
        value = f"{fallback}_{value}"
    return value[:max_len].rstrip("_") or fallback


def _unique_name(base: str, used: set[str], max_len: int = 64) -> str:
    name = _safe_identifier(base, "field", max_len)
    if name not in used:
        used.add(name)
        return name

    suffix = 2
    while True:
        suffix_text = f"_{suffix}"
        candidate = f"{name[:max_len - len(suffix_text)]}{suffix_text}"
        if candidate not in used:
            used.add(candidate)
            return candidate
        suffix += 1


def _safe_choice_value(value: str, fallback: str, used: set[str]) -> str:
    value = html.unescape(str(value or "")).strip()
    value = re.sub(r"\s+", "_", value)
    value = re.sub(r"[^A-Za-z0-9_.:-]+", "_", value).strip("_")
    if not value:
        value = fallback
    if value in used:
        suffix = 2
        base = value[:58]
        while f"{base}_{suffix}" in used:
            suffix += 1
        value = f"{base}_{suffix}"
    used.add(value)
    return value


def _label_key(ref_str: str) -> str:
    # When a ref is not in jr:itext('id') form, the original string is returned
    # and will simply miss the itext lookup, yielding an empty label.
    match = re.search(r"jr:itext\(['\"](.+?)['\"]\)", ref_str or "")
    return match.group(1) if match else ref_str


def _find_itext(root: ET.Element) -> dict[str, dict[str, str]]:
    itext: dict[str, dict[str, str]] = {}
    for trans in root.iter():
        if _strip_ns(trans.tag) != "translation":
            continue
        lang = trans.attrib.get("lang", "en")
        for text_el in trans:
            if _strip_ns(text_el.tag) != "text":
                continue
            tid = text_el.attrib.get("id", "")
            for val in text_el:
                if _strip_ns(val.tag) == "value":
                    itext.setdefault(tid, {})[lang] = (val.text or "").strip()
    return itext


def _find_languages(root: ET.Element) -> list[str]:
    languages: list[str] = []
    for trans in root.iter():
        if _strip_ns(trans.tag) == "translation":
            lang = trans.attrib.get("lang", "en")
            if lang not in languages:
                languages.append(lang)
    return languages or ["en"]


def _find_binds(root: ET.Element) -> tuple[dict[str, dict[str, str]], dict[str, list[dict[str, str]]]]:
    binds_by_path: dict[str, dict[str, str]] = {}
    binds_by_leaf: dict[str, list[dict[str, str]]] = {}

    for bind in root.iter():
        if _strip_ns(bind.tag) != "bind":
            continue
        path = _normalize_ref(bind.attrib.get("nodeset", ""))
        leaf = _ref_leaf(path)
        bind_data = {
            "path": path,
            "leaf": leaf,
            "type": bind.attrib.get("type", ""),
            "required": bind.attrib.get("required", ""),
            "relevant": bind.attrib.get("relevant", ""),
            "calculate": bind.attrib.get("calculate", ""),
        }
        if path:
            binds_by_path[path] = bind_data
        if leaf:
            binds_by_leaf.setdefault(leaf, []).append(bind_data)

    return binds_by_path, binds_by_leaf


def _get_bind(
    ref: str,
    binds_by_path: dict[str, dict[str, str]],
    binds_by_leaf: dict[str, list[dict[str, str]]],
) -> dict[str, str]:
    path = _normalize_ref(ref)
    if path in binds_by_path:
        return binds_by_path[path]

    matches = binds_by_leaf.get(_ref_leaf(ref), [])
    return matches[0] if len(matches) == 1 else {}


def _xsd_to_xlsform_type(xsd_type: str, body_type: str) -> str:
    mapping = {
        "xsd:string": "text",
        "xsd:integer": "integer",
        "xsd:int": "integer",
        "xsd:decimal": "decimal",
        "xsd:date": "date",
        "xsd:time": "time",
        "xsd:dateTime": "dateTime",
        "xsd:boolean": "text",
        "xsd:geopoint": "geopoint",
    }
    if xsd_type in mapping:
        return mapping[xsd_type]
    return {"input": "text", "trigger": "note", "upload": "image"}.get(body_type, "text")


def _resolve_labels(el: ET.Element, itext: dict[str, dict[str, str]], languages: list[str]) -> dict[str, str]:
    labels: dict[str, str] = {}
    for child in el:
        if _strip_ns(child.tag) != "label":
            continue
        ref = child.attrib.get("ref", "")
        if ref:
            tid = _label_key(ref)
            for lang in languages:
                labels[f"label::{lang}"] = itext.get(tid, {}).get(lang, "")
        else:
            text = (child.text or "").strip()
            for lang in languages:
                labels[f"label::{lang}"] = text
        break
    return labels


def _build_choice_list(
    q_name: str,
    el: ET.Element,
    itext: dict[str, dict[str, str]],
    languages: list[str],
    choice_lists: dict[str, list[tuple[str, dict[str, str]]]],
) -> str:
    items: list[tuple[str, dict[str, str]]] = []
    used_values: set[str] = set()

    for index, child in enumerate(el, start=1):
        if _strip_ns(child.tag) != "item":
            continue

        raw_value = ""
        item_labels: dict[str, str] = {}
        for sub in child:
            stag = _strip_ns(sub.tag)
            if stag == "value":
                raw_value = (sub.text or "").strip()
            elif stag == "label":
                ref = sub.attrib.get("ref", "")
                if ref:
                    tid = _label_key(ref)
                    for lang in languages:
                        item_labels[lang] = itext.get(tid, {}).get(lang, "")
                else:
                    for lang in languages:
                        item_labels[lang] = (sub.text or "").strip()

        value = _unique_name(raw_value or f"choice_{index}", used_values, max_len=64)
        items.append((value, item_labels))

    list_name = _safe_identifier(q_name.lower(), "list", max_len=30)
    if list_name in choice_lists:
        existing = choice_lists[list_name]
        if [i[0] for i in existing] == [i[0] for i in items]:
            return list_name
        suffix = 2
        while f"{list_name}_{suffix}" in choice_lists:
            suffix += 1
        list_name = f"{list_name}_{suffix}"

    choice_lists[list_name] = items
    return list_name


def _replace_absolute_refs(expr: str, name_map: dict[str, str], used_names: set[str]) -> str:
    def replace(match: re.Match[str]) -> str:
        raw = match.group(0)
        normalized = _normalize_ref(raw)
        if normalized not in name_map:
            name_map[normalized] = _unique_name(_ref_leaf(normalized), used_names)
        return f"${{{name_map[normalized]}}}"

    return re.sub(r"(?:#form|/data)(?:/[A-Za-z0-9_.:-]+)+", replace, expr)


def _clean_expression(expr: str, name_map: dict[str, str], used_names: set[str]) -> str:
    if not expr:
        return ""
    return _replace_absolute_refs(expr, name_map, used_names)


def _make_body_name(ref: str, used_names: set[str], name_map: dict[str, str]) -> str:
    path = _normalize_ref(ref)
    leaf = _ref_leaf(path)
    if path in name_map:
        return name_map[path]
    name = _unique_name(leaf, used_names)
    if path:
        name_map[path] = name
    return name


def _walk_body(
    el: ET.Element,
    itext: dict[str, dict[str, str]],
    binds_by_path: dict[str, dict[str, str]],
    binds_by_leaf: dict[str, list[dict[str, str]]],
    choice_lists: dict[str, list[tuple[str, dict[str, str]]]],
    survey_rows: list[dict[str, Any]],
    languages: list[str],
    used_names: set[str],
    name_map: dict[str, str],
) -> None:
    tag = _strip_ns(el.tag)

    if tag in ("group", "repeat"):
        ref = el.attrib.get("ref", el.attrib.get("vellum:ref", ""))
        name = _make_body_name(ref, used_names, name_map)
        row = {
            "type": "begin_group" if tag == "group" else "begin_repeat",
            "name": name,
            "appearance": el.attrib.get("appearance", ""),
        }
        row.update(_resolve_labels(el, itext, languages))
        survey_rows.append(row)

        for child in el:
            _walk_body(
                child,
                itext,
                binds_by_path,
                binds_by_leaf,
                choice_lists,
                survey_rows,
                languages,
                used_names,
                name_map,
            )

        survey_rows.append({"type": f"end_{tag}", "name": name})
        return

    if tag not in ("input", "select1", "select", "trigger", "upload"):
        return

    ref = el.attrib.get("ref", el.attrib.get("vellum:ref", ""))
    name = _make_body_name(ref, used_names, name_map)
    bind = _get_bind(ref, binds_by_path, binds_by_leaf)
    xsd_type = bind.get("type", "")

    if tag == "trigger":
        xlstype = "note"
    elif tag in ("select1", "select"):
        list_name = _build_choice_list(name, el, itext, languages, choice_lists)
        prefix = "select_one" if tag == "select1" else "select_multiple"
        xlstype = f"{prefix} {list_name}"
    else:
        xlstype = _xsd_to_xlsform_type(xsd_type, tag)

    row = {
        "type": xlstype,
        "name": name,
        "required": "yes" if bind.get("required") == "true()" else "",
        "relevant": _clean_expression(bind.get("relevant", ""), name_map, used_names),
        "calculation": _clean_expression(bind.get("calculate", ""), name_map, used_names),
    }
    row.update(_resolve_labels(el, itext, languages))
    survey_rows.append(row)

    # CommCare emits a hidden `score_<question>` calculate bind alongside scored
    # questions; pull it in as a calculate row so the scoring logic survives.
    score_path = _normalize_ref(ref).rsplit("/", 1)[0] + f"/score_{_ref_leaf(ref)}"
    score_bind = binds_by_path.get(score_path)
    if score_bind and score_bind.get("calculate"):
        score_name = _make_body_name(score_path, used_names, name_map)
        survey_rows.append({
            "type": "calculate",
            "name": score_name,
            "calculation": _clean_expression(score_bind["calculate"], name_map, used_names),
        })


def parse_xform(xml_str: str) -> dict[str, Any]:
    root = ET.fromstring(xml_str)

    title_el = root.find(f".//{{{HTML_NS}}}title")
    title = (title_el.text or "Untitled Form").strip() if title_el is not None else "Untitled Form"

    data_el = root.find(f".//{{{XFORM_NS}}}instance/*")
    form_id = data_el.attrib.get("name", title) if data_el is not None else title

    itext = _find_itext(root)
    languages = _find_languages(root)
    binds_by_path, binds_by_leaf = _find_binds(root)
    body_el = root.find(f"{{{HTML_NS}}}body")

    survey_rows: list[dict[str, Any]] = []
    choice_lists: dict[str, list[tuple[str, dict[str, str]]]] = {}
    used_names: set[str] = set()
    name_map: dict[str, str] = {}

    if body_el is not None:
        for child in body_el:
            _walk_body(
                child,
                itext,
                binds_by_path,
                binds_by_leaf,
                choice_lists,
                survey_rows,
                languages,
                used_names,
                name_map,
            )

    body_names = {r["name"] for r in survey_rows if "name" in r}
    for path, bind in binds_by_path.items():
        calc = bind.get("calculate", "")
        if calc and name_map.get(path, "") not in body_names:
            calc_name = _make_body_name(path, used_names, name_map)
            survey_rows.append({
                "type": "calculate",
                "name": calc_name,
                "calculation": _clean_expression(calc, name_map, used_names),
            })

    for tid in itext:
        if "thank" in tid.lower():
            labels = {f"label::{lang}": itext[tid].get(lang, "") for lang in languages}
            survey_rows.append({"type": "note", "name": _unique_name("thank_you", used_names), **labels})
            break

    return {
        "title": title,
        "form_id": form_id,
        "languages": languages,
        "survey_rows": survey_rows,
        "choice_lists": choice_lists,
    }


def _translated_text(value: Any, default: str = "") -> str:
    if isinstance(value, dict):
        for text in value.values():
            if text:
                return str(text)
        return default
    return str(value or default)


def _question_languages(questions: list[dict[str, Any]]) -> list[str]:
    languages: list[str] = []
    for question in questions:
        translations = question.get("translations") or {}
        if isinstance(translations, dict):
            for lang in translations:
                if lang not in languages:
                    languages.append(lang)
        for option in question.get("options") or []:
            translations = option.get("translations") or {}
            if isinstance(translations, dict):
                for lang in translations:
                    if lang not in languages:
                        languages.append(lang)
    return languages or ["default"]


def _schema_labels(item: dict[str, Any], languages: list[str], prefix: str = "label") -> dict[str, str]:
    translations = item.get("translations") or {}
    fallback = str(item.get("label") or "")
    labels: dict[str, str] = {}
    for lang in languages:
        if isinstance(translations, dict):
            labels[f"{prefix}::{lang}"] = str(translations.get(lang) or fallback)
        else:
            labels[f"{prefix}::{lang}"] = fallback
    return labels


def _fixture_type_from_data_source(data_source: dict[str, Any]) -> str:
    fixture_type = str(data_source.get("instance_id") or "")
    if fixture_type:
        return fixture_type
    match = re.search(r"item-list:([^'\"/]+)", str(data_source.get("instance_ref") or ""))
    return match.group(1) if match else ""


def _fixture_field_value(row: dict[str, Any], field_name: str) -> str:
    value = (row.get("fields") or {}).get(field_name, "")
    if isinstance(value, dict):
        values = value.get("field_list") or []
        if values:
            return str(values[0].get("field_value") or "")
        return ""
    return str(value or "")


def _choice_extra_column(field_name: str) -> str:
    column = _safe_identifier(field_name, "field", max_len=64)
    reserved = {"list_name", "name"}
    if column in reserved or column.startswith("label") or column.startswith("image"):
        column = _safe_identifier(f"lookup_{column}", "lookup_field", max_len=64)
    return column


def _build_fixture_choice_list(
    question: dict[str, Any],
    name: str,
    choice_lists: dict[str, list[Any]],
    languages: list[str],
    lookup_tables: dict[str, list[dict[str, Any]]],
) -> str:
    data_source = question.get("data_source") or {}
    fixture_type = _fixture_type_from_data_source(data_source)
    rows = lookup_tables.get(fixture_type, [])
    value_ref = str(data_source.get("value_ref") or "id")
    label_ref = str(data_source.get("label_ref") or value_ref)

    list_name = _safe_identifier(name.lower(), "list", max_len=30)
    if list_name in choice_lists:
        suffix = 2
        while f"{list_name}_{suffix}" in choice_lists:
            suffix += 1
        list_name = f"{list_name}_{suffix}"

    used_values: set[str] = set()
    choice_rows: list[dict[str, Any]] = []
    for index, fixture_row in enumerate(rows, start=1):
        fields = fixture_row.get("fields") or {}
        raw_value = _fixture_field_value(fixture_row, value_ref) or str(fixture_row.get("id") or f"choice_{index}")
        raw_label = _fixture_field_value(fixture_row, label_ref) or raw_value
        choice_row: dict[str, Any] = {
            "name": _safe_choice_value(raw_value, f"choice_{index}", used_values),
            "_extra": {},
        }
        for lang in languages:
            choice_row[f"label::{lang}"] = raw_label
        for field_name in fields:
            choice_row["_extra"][_choice_extra_column(field_name)] = _fixture_field_value(fixture_row, field_name)
        choice_rows.append(choice_row)

    choice_lists[list_name] = choice_rows
    return list_name


def _schema_question_type(
    question: dict[str, Any],
    name: str,
    choice_lists: dict[str, list[Any]],
    languages: list[str],
    lookup_tables: dict[str, list[dict[str, Any]]],
) -> str:
    tag = str(question.get("tag") or "").lower()
    qtype = str(question.get("type") or "").lower()

    if tag == "trigger":
        return "note"
    if tag == "hidden" or qtype == "databindonly":
        return "calculate" if question.get("calculate") else "text"
    if tag in {"select1", "select"}:
        data_source = question.get("data_source") or {}
        fixture_type = _fixture_type_from_data_source(data_source)
        if data_source and lookup_tables.get(fixture_type):
            list_name = _build_fixture_choice_list(question, name, choice_lists, languages, lookup_tables)
            return f"{'select_one' if tag == 'select1' else 'select_multiple'} {list_name}"

        if not question.get("options"):
            return "text"

        list_name = _safe_identifier(name.lower(), "list", max_len=30)
        if list_name in choice_lists:
            suffix = 2
            while f"{list_name}_{suffix}" in choice_lists:
                suffix += 1
            list_name = f"{list_name}_{suffix}"

        used_values: set[str] = set()
        options: list[tuple[str, dict[str, str]]] = []
        for index, option in enumerate(question.get("options") or [], start=1):
            value = _unique_name(str(option.get("value") or f"choice_{index}"), used_values)
            option_labels = {
                lang: _schema_labels(option, [lang])[f"label::{lang}"]
                for lang in languages
            }
            options.append((value, option_labels))
        choice_lists[list_name] = options
        return f"{'select_one' if tag == 'select1' else 'select_multiple'} {list_name}"
    if tag == "upload":
        return {"audio": "audio", "image": "image"}.get(qtype, "file")
    if qtype in {"int", "integer"}:
        return "integer"
    if qtype in {"double", "decimal"}:
        return "decimal"
    if qtype == "date":
        return "date"
    if qtype == "datetime":
        return "dateTime"
    if qtype == "geopoint":
        return "geopoint"
    return "text"


def _is_path_within(path: str, parent: str) -> bool:
    path = _normalize_ref(path)
    parent = _normalize_ref(parent)
    return path == parent or path.startswith(f"{parent}/")


def _clean_schema_expression(
    expr: str,
    name_map: dict[str, str],
    used_names: set[str],
    warnings: list[str],
    row_name: str,
    field_name: str,
) -> str:
    expr = str(expr or "")
    if not expr:
        return ""
    if "instance('casedb')" in expr or 'instance("casedb")' in expr:
        warnings.append(f"{row_name}: omitted CommCare case-database {field_name}; Kobo cannot resolve casedb expressions.")
        return ""
    if "instance('commcaresession')" in expr or 'instance("commcaresession")' in expr:
        warnings.append(f"{row_name}: omitted CommCare session {field_name}; Kobo cannot resolve commcaresession expressions.")
        return ""
    if "instance(" in expr:
        warnings.append(f"{row_name}: omitted external-instance {field_name}; Kobo needs the matching external media file.")
        return ""
    return _clean_expression(expr, name_map, used_names)


def _schema_choice_filter(question: dict[str, Any], name_map: dict[str, str]) -> str:
    data_source = question.get("data_source") or {}
    nodeset = str(data_source.get("nodeset") or "")
    match = re.search(r"\[([^][]+)\]", nodeset)
    if not match:
        return ""
    predicate = match.group(1).strip()
    match = re.match(r"([A-Za-z_][\w.-]*)\s*=\s*((?:#form|/data)(?:/[A-Za-z0-9_.:-]+)+)$", predicate)
    if match:
        field_name = _choice_extra_column(match.group(1))
        survey_name = name_map.get(_normalize_ref(match.group(2)), _ref_leaf(match.group(2)))
        return f"{field_name}=${{{survey_name}}}"
    match = re.match(r"((?:#form|/data)(?:/[A-Za-z0-9_.:-]+)+)\s*=\s*([A-Za-z_][\w.-]*)$", predicate)
    if match:
        survey_name = name_map.get(_normalize_ref(match.group(1)), _ref_leaf(match.group(1)))
        field_name = _choice_extra_column(match.group(2))
        return f"{field_name}=${{{survey_name}}}"
    return ""


def parse_commcare_schema(
    app: dict[str, Any],
    module: dict[str, Any],
    form: dict[str, Any],
    lookup_tables: dict[str, list[dict[str, Any]]] | None = None,
) -> dict[str, Any]:
    lookup_tables = lookup_tables or {}
    questions = form.get("questions") or []
    languages = _question_languages(questions)
    app_name = _translated_text(app.get("name"), "CommCare App")
    module_name = _translated_text(module.get("name"), "Module")
    form_name = _translated_text(form.get("name"), "Untitled Form")
    title = f"{app_name} - {module_name} - {form_name}"

    survey_rows: list[dict[str, Any]] = []
    choice_lists: dict[str, list[tuple[str, dict[str, str]]]] = {}
    used_names: set[str] = set()
    name_map: dict[str, str] = {}
    open_groups: list[tuple[str, str, str]] = []
    warnings: list[str] = []

    def close_to(parent_path: str) -> None:
        while open_groups and not _is_path_within(parent_path, open_groups[-1][0]):
            _path, name, tag = open_groups.pop()
            survey_rows.append({"type": f"end_{tag}", "name": name})

    for question in questions:
        path = _normalize_ref(str(question.get("value") or question.get("hashtagValue") or ""))
        tag = str(question.get("tag") or "").lower()
        is_group = bool(question.get("is_group")) or tag in {"group", "repeat"}
        parent_path = _normalize_ref(str(question.get("group") or question.get("repeat") or "/data"))

        close_to(parent_path)

        if is_group:
            name = _make_body_name(path, used_names, name_map)
            group_tag = "repeat" if tag == "repeat" or str(question.get("type") or "").lower() == "repeat" else "group"
            row = {
                "type": "begin_repeat" if group_tag == "repeat" else "begin_group",
                "name": name,
                "relevant": _clean_schema_expression(
                    str(question.get("relevant") or ""),
                    name_map,
                    used_names,
                    warnings,
                    name,
                    "relevance",
                ),
            }
            row.update(_schema_labels(question, languages))
            survey_rows.append(row)
            open_groups.append((path, name, group_tag))
            continue

        name = _make_body_name(path, used_names, name_map)
        row_type = _schema_question_type(question, name, choice_lists, languages, lookup_tables)
        row = {
            "type": row_type,
            "name": name,
            "required": "yes" if question.get("required") else "",
            "relevant": _clean_schema_expression(
                str(question.get("relevant") or ""),
                name_map,
                used_names,
                warnings,
                name,
                "relevance",
            ),
            "calculation": _clean_schema_expression(
                str(question.get("calculate") or question.get("setvalue") or ""),
                name_map,
                used_names,
                warnings,
                name,
                "calculation",
            ),
            "constraint": _clean_schema_expression(
                str(question.get("constraint") or ""),
                name_map,
                used_names,
                warnings,
                name,
                "constraint",
            ),
        }
        choice_filter = _schema_choice_filter(question, name_map)
        if choice_filter:
            row["choice_filter"] = choice_filter
        if row_type == "calculate" and not row["calculation"]:
            row["calculation"] = "''"
        row.update(_schema_labels(question, languages))
        if question.get("comment"):
            for lang in languages:
                row[f"hint::{lang}"] = str(question["comment"])
        if row_type == "text" and str(question.get("tag") or "").lower() in {"select1", "select"} and not question.get("options"):
            data_source = question.get("data_source") or {}
            fixture_type = _fixture_type_from_data_source(data_source)
            if data_source and not lookup_tables.get(fixture_type):
                warnings.append(
                    f"{name}: converted CommCare dynamic data-source select to text; "
                    f"lookup table {fixture_type!r} was not available from the Fixture API."
                )
                for lang in languages:
                    existing_hint = row.get(f"hint::{lang}", "")
                    extra_hint = "CommCare dynamic lookup converted to text."
                    row[f"hint::{lang}"] = f"{existing_hint} {extra_hint}".strip()
            else:
                warnings.append(f"{name}: converted select with no static or lookup-table options to text.")
        survey_rows.append(row)

    close_to("/data")

    form_id = _safe_identifier(f"{app.get('id', app_name)}_{module_name}_{form_name}", "form", max_len=64)
    return {
        "title": title,
        "form_id": form_id,
        "languages": languages,
        "survey_rows": survey_rows,
        "choice_lists": choice_lists,
        "warnings": warnings,
    }


HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
GROUP_FILL = PatternFill("solid", fgColor="D9E1F2")
CALC_FILL = PatternFill("solid", fgColor="E2EFDA")
DATA_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(bottom=THIN, right=THIN)


def _normalize_choice_item(item: Any, languages: list[str]) -> dict[str, Any]:
    """Coerce a choice-list item to a uniform dict.

    Static selects store items as `(value, {lang: label})` tuples; fixture-backed
    selects store them as dicts with `name`, `label::<lang>`, and `_extra` keys.
    """
    if isinstance(item, dict):
        normalized = {"name": item.get("name", ""), "_extra": item.get("_extra") or {}}
        for lang in languages:
            normalized[f"label::{lang}"] = item.get(f"label::{lang}", "")
        return normalized

    value, item_labels = item
    normalized = {"name": value, "_extra": {}}
    for lang in languages:
        normalized[f"label::{lang}"] = item_labels.get(lang, "")
    return normalized


def _write_sheet(ws, headers: list[str], rows: list[dict[str, Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 22

    for row_cells in ws.iter_rows(min_row=2):
        rtype = row_cells[0].value or ""
        for cell in row_cells:
            cell.font = DATA_FONT
            cell.border = CELL_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if "group" in str(rtype).lower() or "repeat" in str(rtype).lower():
            for cell in row_cells:
                cell.fill = GROUP_FILL
        elif rtype == "calculate":
            for cell in row_cells:
                cell.fill = CALC_FILL

    ws.freeze_panes = "A2"


def build_xlsform(parsed: dict[str, Any]) -> Workbook:
    languages = parsed["languages"]
    survey_rows = parsed["survey_rows"]
    choice_lists = parsed["choice_lists"]
    label_cols = [f"label::{lang}" for lang in languages]
    hint_cols = [f"hint::{lang}" for lang in languages]

    survey_headers = (
        ["type", "name"] + label_cols + hint_cols +
        [
            "required",
            "relevant",
            "choice_filter",
            "calculation",
            "appearance",
            "constraint",
            "constraint_message",
            "default",
            "read_only",
        ]
    )

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "survey"
    _write_sheet(ws, survey_headers, survey_rows)

    widths = [20, 35] + [55] * len(label_cols) + [35] * len(hint_cols) + [
        8,
        40,
        35,
        55,
        15,
        40,
        40,
        15,
        10,
    ]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    choices_sheet = wb.create_sheet("choices")
    normalized_lists = {
        list_name: [_normalize_choice_item(item, languages) for item in items]
        for list_name, items in choice_lists.items()
    }
    extra_choice_cols: list[str] = []
    for items in normalized_lists.values():
        for item in items:
            for col in item["_extra"]:
                if col not in extra_choice_cols:
                    extra_choice_cols.append(col)
    choices_headers = ["list_name", "name"] + label_cols + extra_choice_cols
    choices_rows: list[dict[str, Any]] = []
    for list_name, items in normalized_lists.items():
        for item in items:
            row = {"list_name": list_name, "name": item["name"]}
            for lang in languages:
                row[f"label::{lang}"] = item[f"label::{lang}"]
            for col, value in item["_extra"].items():
                row[col] = value
            choices_rows.append(row)
    _write_sheet(choices_sheet, choices_headers, choices_rows)
    for idx, width in enumerate([20, 30] + [50] * len(label_cols) + [24] * len(extra_choice_cols), 1):
        choices_sheet.column_dimensions[get_column_letter(idx)].width = width

    settings_sheet = wb.create_sheet("settings")
    settings_headers = ["form_title", "form_id", "version", "default_language", "instance_name"]
    settings_data = {
        "form_title": parsed["title"],
        "form_id": _safe_identifier(parsed["form_id"].lower(), "form", max_len=64),
        "version": "1",
        "default_language": languages[0],
        "instance_name": "",
    }
    _write_sheet(settings_sheet, settings_headers, [settings_data])
    for idx, width in enumerate([40, 40, 10, 25, 40], 1):
        settings_sheet.column_dimensions[get_column_letter(idx)].width = width

    return wb


def validate_xlsform(parsed: dict[str, Any]) -> list[str]:
    warnings: list[str] = list(parsed.get("warnings", []))
    rows = parsed["survey_rows"]
    choice_lists = parsed["choice_lists"]
    names = [row.get("name", "") for row in rows if row.get("name")]
    counts = Counter(names)

    for name, count in sorted(counts.items()):
        if count > 1:
            types = [row.get("type", "") for row in rows if row.get("name") == name]
            if not all(t.startswith("end_") or t.startswith("begin_") for t in types):
                warnings.append(f"duplicate survey name: {name}")

    for index, row in enumerate(rows, start=2):
        name = row.get("name", "")
        row_type = row.get("type", "")
        if not name:
            warnings.append(f"survey row {index} has a blank name")
        elif not VALID_XLSFORM_NAME.match(name):
            warnings.append(f"survey row {index} has invalid name: {name}")
        if row_type.startswith("select_"):
            list_name = row_type.split(maxsplit=1)[1] if " " in row_type else ""
            if list_name not in choice_lists:
                warnings.append(f"survey row {index} references missing choice list: {list_name}")

    for list_name, items in choice_lists.items():
        seen: set[str] = set()
        for item in items:
            value = str(item.get("name", "")) if isinstance(item, dict) else str(item[0])
            if not VALID_CHOICE_NAME.match(value):
                warnings.append(f"choice list {list_name} has invalid choice name: {value}")
            if value in seen:
                warnings.append(f"choice list {list_name} has duplicate choice name: {value}")
            seen.add(value)

    return warnings


def workbook_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _response_summary(resp: requests.Response | None) -> str:
    if resp is None:
        return ""
    try:
        return str(resp.json())
    except ValueError:
        return resp.text[:1000]


def _kobo_session(token: str) -> requests.Session:
    session = requests.Session()
    session.headers.update({"Authorization": f"Token {token}"})
    return session


def kobo_upload_xlsform(session: requests.Session, server_url: str, xlsx_bytes: bytes, filename: str) -> dict:
    base = server_url.rstrip("/")

    create_resp = session.post(
        f"{base}/api/v2/assets/?format=json",
        json={"name": filename.rsplit(".", 1)[0], "asset_type": "survey"},
        timeout=30,
    )
    create_resp.raise_for_status()
    asset_uid = create_resp.json()["uid"]

    import_resp = session.post(
        f"{base}/api/v2/imports/?format=json",
        data={"library": "false", "destination": f"{base}/api/v2/assets/{asset_uid}/"},
        files={
            "file": (
                filename,
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        timeout=60,
    )
    import_resp.raise_for_status()
    poll_url = import_resp.json().get("url")
    if not poll_url:
        raise ValueError(f"Kobo import response did not include a poll URL: {import_resp.text}")

    poll_attempts = 30
    poll_interval = 2
    for _ in range(poll_attempts):
        time.sleep(poll_interval)
        poll = session.get(poll_url, timeout=30)
        poll.raise_for_status()
        result = poll.json()
        if result.get("status") == "complete":
            updated = result.get("messages", {}).get("updated", [])
            if updated:
                return updated[0]
            raise ValueError(f"Import complete but no asset was updated: {result}")
        if result.get("status") == "error":
            raise ValueError(f"Import failed: {result}")

    raise TimeoutError(
        f"Import did not complete after {poll_attempts * poll_interval} seconds"
    )


def kobo_deploy_asset(session: requests.Session, server_url: str, asset_uid: str) -> dict:
    url = f"{server_url.rstrip('/')}/api/v2/assets/{asset_uid}/deployment/"
    resp = session.post(url, data={"active": "true"}, timeout=60)
    resp.raise_for_status()
    return resp.json()


def _extract_xform_from_word_xml(raw: str) -> str:
    texts = re.findall(r"<w:t[^>]*>([^<]+)</w:t>", raw)
    unescaped = html.unescape("".join(texts))
    start = unescaped.find("<h:html")
    if start == -1:
        start = unescaped.find("<?xml")
    if start == -1:
        raise ValueError("No XForm XML found inside Word document")
    return unescaped[start:]


def _load_xml_file(path: Path) -> str:
    raw = path.read_text(encoding="utf-8")
    if "pkg:package" in raw or "mso-application" in raw:
        return _extract_xform_from_word_xml(raw)
    return raw


def safe_filename(title: str) -> str:
    return _safe_identifier(title, "form", max_len=80)


def _collect_sources(args: argparse.Namespace) -> list[tuple[str, str | dict[str, Any]]]:
    if args.commcare_fetch:
        if not args.commcare_domain or not args.commcare_user or not args.commcare_token:
            raise ValueError("CommCare fetch mode requires domain, user, and token.")
        session = _commcare_session(args.commcare_user, args.commcare_token)
        apps = _list_commcare_apps(session, args.commcare_domain, args.commcare_limit)
        fixture_types = _collect_fixture_types(apps)
        lookup_tables: dict[str, list[dict[str, Any]]] = {}
        for fixture_type in fixture_types:
            lookup_tables[fixture_type] = _list_commcare_fixture_rows(session, args.commcare_domain, fixture_type)
        if lookup_tables:
            print(
                "  Fetched lookup table rows: "
                + ", ".join(f"{name}={len(rows)}" for name, rows in sorted(lookup_tables.items()))
            )
        sources: list[tuple[str, dict[str, Any]]] = []
        for app in apps:
            for module in app.get("modules", []):
                for form in module.get("forms", []):
                    parsed = parse_commcare_schema(app, module, form, lookup_tables)
                    sources.append((parsed["title"], parsed))

        print(f"  Found {len(apps)} app(s) and {len(sources)} form schema(s) from CommCare\n")
        return sources

    folder = Path(args.input_folder)
    if not folder.exists():
        raise FileNotFoundError(f"Input folder not found: {folder}")
    xml_files = sorted(folder.glob("*.xml"))
    if not xml_files:
        raise FileNotFoundError(f"No .xml files found in: {folder}")

    print(f"  Found {len(xml_files)} .xml file(s) in {folder}\n")
    sources = []
    for path in xml_files:
        try:
            sources.append((path.stem, _load_xml_file(path)))
        except Exception as exc:
            print(f"  [ERR ] {path.name}: {exc}")
    return sources


def _process_source(
    label: str,
    source: str | dict[str, Any],
    out_dir: Path | None,
) -> tuple[dict[str, Any] | None, Workbook | None, str | None, list[str]]:
    parsed = source if isinstance(source, dict) else parse_xform(source)
    warnings = validate_xlsform(parsed)
    wb = build_xlsform(parsed)
    filename = f"{safe_filename(parsed['title'])}.xlsx"
    if out_dir is not None:
        out_dir.mkdir(parents=True, exist_ok=True)
        wb.save(out_dir / filename)
    return parsed, wb, filename, warnings


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Convert CommCare XForm XML files to Kobo XLSForm.")
    parser.add_argument("--input-folder", default=str(XML_INPUT_FOLDER), help="Folder containing .xml files.")
    parser.add_argument("--output-folder", default=str(XLSFORM_OUTPUT_FOLDER), help="Folder for generated .xlsx files.")
    parser.add_argument(
        "--save",
        action=argparse.BooleanOptionalAction,
        default=SAVE_XLSFORMS_LOCALLY,
        help="Save generated .xlsx files locally.",
    )
    parser.add_argument(
        "--upload",
        action=argparse.BooleanOptionalAction,
        default=UPLOAD_TO_KOBO,
        help="Upload generated XLSForms to KoboToolbox.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Validate only; do not upload even if UPLOAD_TO_KOBO is true.")
    parser.add_argument(
        "--deploy",
        action=argparse.BooleanOptionalAction,
        default=KOBO_DEPLOY,
        help="Deploy uploaded Kobo forms.",
    )
    parser.add_argument("--kobo-server-url", default=KOBO_SERVER_URL, help="KoboToolbox server URL.")
    parser.add_argument("--kobo-api-token", default=KOBO_API_TOKEN, help="Kobo API token. Prefer KOBO_API_TOKEN env var.")
    parser.add_argument(
        "--commcare-fetch",
        action=argparse.BooleanOptionalAction,
        default=COMMCARE_FETCH,
        help="Fetch CommCare app/form schema instead of reading XML_INPUT_FOLDER.",
    )
    parser.add_argument("--commcare-domain", default=COMMCARE_DOMAIN, help="CommCare project domain.")
    parser.add_argument("--commcare-user", default=COMMCARE_USER, help="CommCare API username/email.")
    parser.add_argument("--commcare-token", default=COMMCARE_TOKEN, help="CommCare API token.")
    parser.add_argument("--commcare-limit", type=int, default=COMMCARE_LIMIT, help="Max CommCare forms to fetch; 0 means all.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_arg_parser().parse_args(argv)
    uploading = bool(args.upload) and not args.dry_run
    save_locally = bool(args.save)

    if uploading and not args.kobo_api_token:
        print("[ERROR] --upload requires KOBO_API_TOKEN or --kobo-api-token.", file=sys.stderr)
        return 2

    print()
    print("-" * 60)
    mode = "CommCare API" if args.commcare_fetch else f"Folder: {args.input_folder}"
    print(f"  Source  : {mode}")
    print(f"  Kobo    : {args.kobo_server_url}")
    print(f"  Upload  : {'YES' if uploading else 'NO'}")
    print(f"  Deploy  : {'YES' if uploading and args.deploy else 'NO'}")
    print(f"  Save XLS: {'YES' if save_locally else 'NO'}")
    print("-" * 60)
    print()

    try:
        sources = _collect_sources(args)
    except Exception as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 1

    out_dir = Path(args.output_folder) if save_locally else None
    kobo_session = _kobo_session(args.kobo_api_token) if uploading else None
    ok = err = warn_count = 0

    for label, xml_str in tqdm(sources, desc="Processing", unit="form"):
        try:
            parsed, wb, filename, warnings = _process_source(label, xml_str, out_dir)
        except ET.ParseError as exc:
            tqdm.write(f"  [ERR ] {label}: XML parse error - {exc}")
            err += 1
            continue
        except Exception as exc:
            tqdm.write(f"  [ERR ] {label}: conversion error - {exc}")
            err += 1
            continue

        warn_count += len(warnings)
        status = "OK" if not warnings else f"OK with {len(warnings)} warning(s)"
        tqdm.write(
            f"  [VAL ] {filename}: {status}; "
            f"{len(parsed['survey_rows'])} survey rows, {len(parsed['choice_lists'])} choice lists"
        )
        for warning in warnings[:5]:
            tqdm.write(f"        warning: {warning}")
        if len(warnings) > 5:
            tqdm.write(f"        ... {len(warnings) - 5} more warning(s)")

        if not uploading:
            ok += 1
            continue

        assert kobo_session is not None
        assert wb is not None
        try:
            asset = kobo_upload_xlsform(
                kobo_session,
                args.kobo_server_url,
                workbook_to_bytes(wb),
                filename or "form.xlsx",
            )
            asset_uid = asset.get("uid", "")
            tqdm.write(f"  [UP  ] {filename} -> uid={asset_uid}")
        except requests.HTTPError as exc:
            body: Any = ""
            try:
                body = exc.response.json()
            except Exception:
                body = exc.response.text if exc.response is not None else ""
            status_code = exc.response.status_code if exc.response is not None else "unknown"
            tqdm.write(f"  [ERR ] {label}: upload failed HTTP {status_code} - {body}")
            err += 1
            continue
        except Exception as exc:
            tqdm.write(f"  [ERR ] {label}: upload error - {exc}")
            err += 1
            continue

        if args.deploy and asset_uid:
            try:
                kobo_deploy_asset(kobo_session, args.kobo_server_url, asset_uid)
                tqdm.write(f"  [DEP ] {filename} deployed")
            except requests.HTTPError as exc:
                status_code = exc.response.status_code if exc.response is not None else "unknown"
                body = _response_summary(exc.response)
                tqdm.write(f"  [WARN] {filename}: deploy failed HTTP {status_code} - {body}")
            except Exception as exc:
                tqdm.write(f"  [WARN] {filename}: deploy error - {exc}")

        ok += 1

    print()
    print("-" * 60)
    action = "uploaded" if uploading else "validated"
    print(f"  Done. {ok} {action}, {err} errors, {warn_count} warnings")
    if save_locally:
        print(f"  XLSForms saved to: {Path(args.output_folder).resolve()}")
    print("-" * 60)
    print()
    return 0 if err == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
